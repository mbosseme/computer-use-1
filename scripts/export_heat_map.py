import pandas as pd
from google.cloud import bigquery
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# BigQuery Setup
client = bigquery.Client()
project_id = "matthew-bossemeyer"
dataset_id = "dataform_hciq"
run_id = "2026-03-04__portfolio-competitiveness"
output_dir = f"runs/{run_id}/"
os.makedirs(output_dir, exist_ok=True)
out_file = os.path.join(output_dir, "Contract_Competitive_Heat_Map_All_Programs_and_Categories.xlsx")

query = f"""
WITH expiration_dates AS (
    SELECT 
        Contract_Number,
        MAX(CAST(Contract_Expiration_Date AS DATE)) as Expiration_Date,
        MAX(service_line) as Service_Line
    FROM `abi-inbound-prod.abi_inbound_bq_stg_purchasing_provider_transaction.transaction_analysis_expanded`
    GROUP BY 1
),
item_gaps AS (
    SELECT 
        c.*,
        e.Expiration_Date,
        e.Service_Line,
        CASE
            WHEN c.Portfolio_Prefix = 'SP' THEN 'Surpass'
            WHEN c.Portfolio_Prefix = 'AD' THEN 'Ascend Drive'
            ELSE 'National'
        END as Program,
        -- The target price based on Joe's rule
        CASE
            WHEN c.Portfolio_Prefix = 'SP' THEN c.hciq_low_benchmark
            WHEN c.Portfolio_Prefix = 'AD' THEN c.hciq_90_benchmark
            ELSE c.hciq_75_benchmark
        END as Target_Price_Per_Unit
    FROM `{project_id}.{dataset_id}.contract_item_benchmark_summary` c
    LEFT JOIN expiration_dates e ON c.Contract_Number = e.Contract_Number
    WHERE 
        c.is_benchmarked = TRUE
        AND e.Expiration_Date BETWEEN '2026-07-01' AND '2027-06-30'
        AND c.Contract_Name != 'UNKNOWN'
),
item_opportunities AS (
    SELECT
        *,
        -- Only positive dollar opportunity (where current contract best price > target price)
        CASE 
            WHEN contract_best_price > Target_Price_Per_Unit AND Target_Price_Per_Unit > 0 
            THEN (contract_best_price - Target_Price_Per_Unit) * Total_Units_6mo
            ELSE 0 
        END as Six_Month_Opportunity
    FROM item_gaps
),
contract_aggregates AS (
    SELECT
        Contract_Name,
        Program,
        ARRAY_AGG(Contracted_Supplier ORDER BY Total_Spend_6mo DESC LIMIT 1)[OFFSET(0)] as Contracted_Supplier,
        MAX(Service_Line) as Service_Line,
        MAX(Expiration_Date) as Expiration_Date,
        COUNT(DISTINCT Manufacturer_Catalog_Number) as Benchmarked_Items,
        SUM(Total_Spend_6mo) as Agg_Total_Spend_6mo,
        SUM(Spend_at_Best_Tier) as Agg_Spend_at_Best_Tier,
        SUM(Six_Month_Opportunity) as Agg_Six_Month_Opportunity,
        SAFE_DIVIDE(SUM(Six_Month_Opportunity), SUM(Spend_at_Best_Tier)) as Opportunity_Pct
    FROM item_opportunities
    GROUP BY Contract_Name, Program
    HAVING Agg_Spend_at_Best_Tier > 0
),
ranked_contracts AS (
    SELECT 
        Contract_Name,
        Program,
        Contracted_Supplier,
        Service_Line,
        Expiration_Date,
        Benchmarked_Items,
        Agg_Total_Spend_6mo * 2 as Annualized_Total_Spend,
        Agg_Spend_at_Best_Tier * 2 as Annualized_Benchmarked_Spend,
        Agg_Six_Month_Opportunity * 2 as Annualized_Savings_Opportunity,
        Opportunity_Pct
    FROM contract_aggregates
    ORDER BY Annualized_Savings_Opportunity DESC
)
SELECT * FROM ranked_contracts
"""

print("Executing BigQuery...")
df = client.query(query).to_dataframe()

print(f"Number of rows fetched: {len(df)}")
if len(df) > 0:
    # Save to Excel
    writer = pd.ExcelWriter(out_file, engine='openpyxl')
    
    # We will write the full heat map (this is the 3rd attachment the user mentioned)
    df.to_excel(writer, sheet_name='All Opportunities', index=False)
    
    writer.close()
    
    # Openpyxl formatting
    wb = openpyxl.load_workbook(out_file)
    ws = wb['All Opportunities']
    
    currency_cols = ['G', 'H', 'I']
    pct_col = 'J'
    date_col = 'E'
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter in currency_cols:
                if cell.value is not None:
                    cell.number_format = '$#,##0'
            elif col_letter == pct_col:
                if cell.value is not None:
                    cell.number_format = '0.0%'
            elif col_letter == date_col:
                if cell.value is not None:
                    cell.number_format = 'yyyy-mm-dd'
                    
    # Heat map based on Savings Opportunity (I)
    ws.conditional_formatting.add(f'I2:I{ws.max_row}', ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='63BE7B'))
    # Heat map based on Opportunity Pct (J)
    ws.conditional_formatting.add(f'J2:J{ws.max_row}', ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='F8696B'))
                                  
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 20

    wb.save(out_file)
    print(f"Exported to {out_file}")
else:
    print("No data found!")
